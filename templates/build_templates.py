#!/usr/bin/env python3
"""Build all 3 teach4texas Excel templates."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from copy import copy
import random
import os

OUT = os.path.dirname(os.path.abspath(__file__))

# ── Shared styling helpers ──────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
ALT_ROW = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
LIGHT_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
LIGHT_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
LIGHT_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="1F4E79")
INST_FONT = Font(name="Calibri", size=11)
INST_BOLD = Font(name="Calibri", size=11, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def style_header(ws, row, ncols, fill=None):
    f = fill or HEADER_FILL
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = f
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_data_rows(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if (r - start_row) % 2 == 1:
                cell.fill = ALT_ROW

def auto_width(ws, ncols, min_w=12, max_w=30):
    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = min(max(min_w, 15), max_w)

def write_instructions(ws, lines):
    """Write instruction lines; lines is list of (text, bold?)."""
    ws.sheet_properties.tabColor = "FFC000"
    ws.column_dimensions["A"].width = 100
    for i, (txt, bold) in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=txt)
        cell.font = INST_BOLD if bold else INST_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")


# ════════════════════════════════════════════════════════════════════════
# 1. STAAR Prep Tracker
# ════════════════════════════════════════════════════════════════════════

def build_staar():
    wb = openpyxl.Workbook()

    # ── Student Roster ──
    ws = wb.active
    ws.title = "Student Roster"
    ws.sheet_properties.tabColor = "1F4E79"
    headers = ["Student Name", "Grade", "Subject", "Pre-Test Score", "Post-Test Score",
               "Growth %", "Meets Prediction", "Masters Prediction", "Notes"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    first_names = ["Maria", "Jose", "Emily", "David", "Sofia", "Carlos", "Aiden", "Mia",
                   "Jayden", "Isabella", "Ethan", "Olivia", "Liam", "Ava", "Noah",
                   "Emma", "Lucas", "Camila", "Mason", "Luna", "Diego", "Valentina",
                   "Elijah", "Ariana", "Mateo"]
    last_names = ["Garcia", "Rodriguez", "Martinez", "Hernandez", "Lopez", "Gonzalez",
                  "Wilson", "Anderson", "Thomas", "Jackson", "White", "Harris", "Martin",
                  "Thompson", "Moore", "Taylor", "Brown", "Davis", "Miller", "Johnson",
                  "Reyes", "Cruz", "Morales", "Ortiz", "Gutierrez"]
    subjects = ["Math", "Reading", "Science", "Social Studies"]
    random.seed(42)
    for i in range(25):
        r = i + 2
        name = f"{first_names[i]}, {last_names[i]}"
        grade = random.choice([3,4,5,6,7,8])
        subj = subjects[i % 4]
        pre = random.randint(30, 75)
        post = min(100, pre + random.randint(5, 30))
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=grade)
        ws.cell(row=r, column=3, value=subj)
        ws.cell(row=r, column=4, value=pre)
        ws.cell(row=r, column=5, value=post)
        ws.cell(row=r, column=6).value = f"=IF(D{r}=0,\"\",ROUND((E{r}-D{r})/D{r}*100,1))"
        ws.cell(row=r, column=6).number_format = '0.0"%"'
        ws.cell(row=r, column=7).value = f'=IF(E{r}>=50,"Likely Meets","At Risk")'
        ws.cell(row=r, column=8).value = f'=IF(E{r}>=75,"Likely Masters",IF(E{r}>=60,"Possible","Unlikely"))'
        notes = ["Strong improvement", "Needs reteach on fractions", "ESL accommodation",
                 "GT student", "504 plan", "Consistent effort", "Tutoring recommended",
                 "Parent conference scheduled", "", ""]
        ws.cell(row=r, column=9, value=random.choice(notes))
    style_data_rows(ws, 2, 26, len(headers))
    auto_width(ws, len(headers))
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["I"].width = 28
    ws.auto_filter.ref = f"A1:I26"

    # ── Class Overview ──
    ws2 = wb.create_sheet("Class Overview")
    ws2.sheet_properties.tabColor = "2E75B6"
    # Summary section
    ws2.cell(row=1, column=1, value="CLASS OVERVIEW DASHBOARD").font = TITLE_FONT
    ws2.merge_cells("A1:E1")

    sum_headers = ["Subject", "Avg Pre-Test", "Avg Post-Test", "Avg Growth %", "Pass Rate (≥50)"]
    for c, h in enumerate(sum_headers, 1):
        ws2.cell(row=3, column=c, value=h)
    style_header(ws2, 3, len(sum_headers))

    for i, subj in enumerate(subjects):
        r = 4 + i
        ws2.cell(row=r, column=1, value=subj)
        ws2.cell(row=r, column=2).value = f'=AVERAGEIF(\'Student Roster\'!C:C,A{r},\'Student Roster\'!D:D)'
        ws2.cell(row=r, column=2).number_format = '0.0'
        ws2.cell(row=r, column=3).value = f'=AVERAGEIF(\'Student Roster\'!C:C,A{r},\'Student Roster\'!E:E)'
        ws2.cell(row=r, column=3).number_format = '0.0'
        ws2.cell(row=r, column=4).value = f'=IF(B{r}=0,"",ROUND((C{r}-B{r})/B{r}*100,1))'
        ws2.cell(row=r, column=4).number_format = '0.0"%"'
        ws2.cell(row=r, column=5).value = f'=COUNTIFS(\'Student Roster\'!C:C,A{r},\'Student Roster\'!E:E,">="&50)/COUNTIF(\'Student Roster\'!C:C,A{r})'
        ws2.cell(row=r, column=5).number_format = '0%'
    style_data_rows(ws2, 4, 7, len(sum_headers))

    # Students at risk
    ws2.cell(row=10, column=1, value="STUDENTS AT RISK (Post-Test < 50)").font = INST_BOLD
    ws2.merge_cells("A10:E10")
    risk_headers = ["Student", "Subject", "Post-Test Score", "Growth %", "Action Needed"]
    for c, h in enumerate(risk_headers, 1):
        ws2.cell(row=11, column=c, value=h)
    style_header(ws2, 11, len(risk_headers))
    ws2.cell(row=12, column=1, value="(Filter Student Roster for Post-Test < 50)")
    auto_width(ws2, 5)

    # ── Weekly Planning ──
    ws3 = wb.create_sheet("Weekly Planning")
    ws3.sheet_properties.tabColor = "548235"
    headers3 = ["Week #", "Date Range", "TEKS Covered", "Practice Test Avg", "Reteach Topics",
                "Small Group Assignments", "Notes"]
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=c, value=h)
    style_header(ws3, 1, len(headers3))

    weeks_data = [
        [1, "Jan 6-10", "3.4A, 3.4B", 52, "Place value, number lines", "Group A: Fractions / Group B: Multiplication", "Baseline week"],
        [2, "Jan 13-17", "3.5A, 3.5B", 58, "Addition strategies", "Group A: Word problems / Group B: Mental math", ""],
        [3, "Jan 20-24", "3.6A, 3.6C", 55, "Geometry basics", "Group A: Shapes / Group B: Area", "MLK holiday Mon"],
        [4, "Jan 27-31", "3.7A, 3.7B", 61, "Measurement", "Group A: Customary / Group B: Metric", ""],
        [5, "Feb 3-7", "3.8A, 3.8B", 63, "Data analysis", "Group A: Graphs / Group B: Tables", "Progress reports"],
        [6, "Feb 10-14", "Review 3.4-3.8", 67, "Mixed review", "Targeted by weakness", "Mock STAAR"],
        [7, "Feb 17-21", "4.2A, 4.2B", 59, "Comparing numbers", "Group A: Decimals / Group B: Fractions", ""],
        [8, "Feb 24-28", "4.3A, 4.3C", 64, "Operations", "Differentiated practice", "Parent night Thurs"],
    ]
    for i, row_data in enumerate(weeks_data):
        r = i + 2
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)
    style_data_rows(ws3, 2, 9, len(headers3))
    auto_width(ws3, len(headers3))
    ws3.column_dimensions["C"].width = 20
    ws3.column_dimensions["E"].width = 25
    ws3.column_dimensions["F"].width = 35

    # ── Instructions ──
    ws4 = wb.create_sheet("Instructions")
    write_instructions(ws4, [
        ("📋 STAAR Prep Tracker — Instructions", True),
        ("", False),
        ("STUDENT ROSTER TAB", True),
        ("• Enter each student's name, grade level, and STAAR subject", False),
        ("• Input Pre-Test scores from your baseline assessment", False),
        ("• Update Post-Test scores after each practice test or benchmark", False),
        ("• Growth %, Meets Prediction, and Masters Prediction auto-calculate", False),
        ("• Meets threshold: 50+ on scale score | Masters threshold: 75+", False),
        ("• Use Notes column for accommodations, parent contacts, interventions", False),
        ("• Use filters (dropdown arrows) to sort by subject, grade, or risk level", False),
        ("", False),
        ("CLASS OVERVIEW TAB", True),
        ("• Automatically pulls averages from Student Roster — no manual entry needed", False),
        ("• Shows pass rate by subject to identify weak areas", False),
        ("• Use 'Students at Risk' section to flag students needing intervention", False),
        ("", False),
        ("WEEKLY PLANNING TAB", True),
        ("• Plan STAAR prep week by week with specific TEKS alignment", False),
        ("• Track practice test averages to monitor class progress over time", False),
        ("• Assign small groups based on reteach needs", False),
        ("• Great for PLC meetings and admin walkthroughs", False),
        ("", False),
        ("TIPS FOR TEXAS TEACHERS", True),
        ("• Align practice tests to released STAAR items (tea.texas.gov)", False),
        ("• Update scores weekly for the most actionable data", False),
        ("• Share Class Overview with your admin to show data-driven instruction", False),
        ("• Use this tracker as evidence for your T-TESS goal-setting conference", False),
        ("", False),
        ("© teach4texas | For personal classroom use only", False),
    ])

    wb.save(os.path.join(OUT, "STAAR_Prep_Tracker.xlsx"))
    print("✅ STAAR_Prep_Tracker.xlsx")


# ════════════════════════════════════════════════════════════════════════
# 2. TIA Portfolio Builder
# ════════════════════════════════════════════════════════════════════════

def build_tia():
    wb = openpyxl.Workbook()

    # ── Observation Tracker ──
    ws = wb.active
    ws.title = "Observation Tracker"
    ws.sheet_properties.tabColor = "7030A0"
    headers = ["Date", "Observer", "Domain", "Dimension", "Score (1-5)", "Notes", "Evidence Summary"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    obs_data = [
        ["2025-09-15", "Dr. Smith (Principal)", "Planning", "1.1 - Standards Focus", 4, "Strong TEKS alignment", "Lesson plan showed vertical alignment to STAAR"],
        ["2025-09-15", "Dr. Smith (Principal)", "Planning", "1.2 - Data & Assessment", 3, "Room for growth in formative assessment", "Used exit tickets but not differentiated"],
        ["2025-10-20", "Ms. Johnson (AP)", "Instruction", "2.1 - Achieving Expectations", 4, "High engagement observed", "95% students on task, strategic questioning"],
        ["2025-10-20", "Ms. Johnson (AP)", "Instruction", "2.3 - Content Knowledge", 5, "Exemplary content delivery", "Cross-curricular connections, real-world examples"],
        ["2025-11-18", "Dr. Smith (Principal)", "Learning Environment", "3.1 - Classroom Culture", 4, "Positive rapport with students", "Restorative practices evident"],
        ["2025-11-18", "Dr. Smith (Principal)", "Learning Environment", "3.3 - Procedures", 5, "Seamless transitions", "Students self-managed materials and groups"],
        ["2026-01-22", "Ms. Johnson (AP)", "Professional Practices", "4.1 - Professional Demeanor", 4, "Team leader behaviors", "Led PLC, mentored new teacher"],
        ["2026-01-22", "Ms. Johnson (AP)", "Instruction", "2.2 - Content Delivery", 4, "Differentiated stations", "3 levels of scaffolded practice observed"],
    ]
    for i, row in enumerate(obs_data):
        r = i + 2
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    style_data_rows(ws, 2, 9, len(headers))
    auto_width(ws, len(headers))
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 40

    # ── Student Growth ──
    ws2 = wb.create_sheet("Student Growth")
    ws2.sheet_properties.tabColor = "548235"
    headers2 = ["Assessment", "Subject", "Baseline (BOY)", "Target", "Actual (EOY)", "Growth %", "Met Target?", "TIA Growth Level"]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, len(headers2))

    growth_data = [
        ["STAAR 2025", "Math", 45, 65, 72, None, None, None],
        ["STAAR 2025", "Reading", 52, 68, 70, None, None, None],
        ["MAP Fall→Spring", "Math", 205, 215, 220, None, None, None],
        ["MAP Fall→Spring", "Reading", 198, 208, 211, None, None, None],
        ["District Benchmark Q1→Q4", "Math", 55, 75, 78, None, None, None],
        ["District Benchmark Q1→Q4", "Reading", 60, 78, 74, None, None, None],
    ]
    for i, row in enumerate(growth_data):
        r = i + 2
        for c, val in enumerate(row, 1):
            if val is not None:
                ws2.cell(row=r, column=c, value=val)
        ws2.cell(row=r, column=6).value = f'=IF(C{r}=0,"",ROUND((E{r}-C{r})/C{r}*100,1))'
        ws2.cell(row=r, column=6).number_format = '0.0"%"'
        ws2.cell(row=r, column=7).value = f'=IF(E{r}>=D{r},"✅ Yes","❌ No")'
        ws2.cell(row=r, column=8).value = f'=IF(F{r}>=40,"Top Quartile",IF(F{r}>=25,"Above Average",IF(F{r}>=10,"Average","Below Average")))'
    style_data_rows(ws2, 2, 7, len(headers2))
    auto_width(ws2, len(headers2))

    # ── PD Log ──
    ws3 = wb.create_sheet("Professional Development")
    ws3.sheet_properties.tabColor = "BF8F00"
    headers3 = ["Date", "Activity", "Hours", "Domain Aligned To", "Certificate/Proof", "Notes"]
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=c, value=h)
    style_header(ws3, 1, len(headers3))

    pd_data = [
        ["2025-08-05", "District New Teacher Academy", 16, "All Domains", "Certificate on file", "Required PD"],
        ["2025-09-12", "Data-Driven Instruction Workshop", 6, "Planning (Domain 1)", "Sign-in sheet", "Region 4 ESC"],
        ["2025-10-15", "Kagan Cooperative Learning", 8, "Instruction (Domain 2)", "Certificate", "Campus PD day"],
        ["2025-11-02", "Restorative Practices Training", 4, "Learning Environment (Domain 3)", "Certificate", "Online course"],
        ["2025-12-10", "PLC Leadership Seminar", 3, "Professional Practices (Domain 4)", "Agenda + notes", "District-led"],
        ["2026-01-18", "STAAR Prep Strategies", 6, "Planning (Domain 1)", "Sign-in sheet", "Region 6 ESC"],
        ["2026-02-08", "Technology Integration (Canvas LMS)", 4, "Instruction (Domain 2)", "Badge earned", "Self-paced"],
    ]
    for i, row in enumerate(pd_data):
        r = i + 2
        for c, val in enumerate(row, 1):
            ws3.cell(row=r, column=c, value=val)
    # Total hours
    r_tot = len(pd_data) + 3
    ws3.cell(row=r_tot, column=2, value="TOTAL HOURS:").font = Font(bold=True)
    ws3.cell(row=r_tot, column=3).value = f"=SUM(C2:C{len(pd_data)+1})"
    ws3.cell(row=r_tot, column=3).font = Font(bold=True)
    style_data_rows(ws3, 2, len(pd_data)+1, len(headers3))
    auto_width(ws3, len(headers3))
    ws3.column_dimensions["B"].width = 35
    ws3.column_dimensions["E"].width = 22

    # ── TIA Designation Calculator ──
    ws4 = wb.create_sheet("TIA Calculator")
    ws4.sheet_properties.tabColor = "FF0000"
    ws4.cell(row=1, column=1, value="TIA DESIGNATION CALCULATOR").font = TITLE_FONT
    ws4.merge_cells("A1:E1")

    ws4.cell(row=3, column=1, value="YOUR SCORES").font = INST_BOLD
    calc_labels = ["Teacher Observation Average (1-5):", "Student Growth Score (0-100):",
                   "Weighted Observation (70%):", "Weighted Growth (30%):", "COMPOSITE SCORE:",
                   "", "PROJECTED DESIGNATION:"]
    for i, lbl in enumerate(calc_labels):
        ws4.cell(row=4+i, column=1, value=lbl).font = INST_BOLD if i in [4,6] else INST_FONT

    # Input cells with formatting
    ws4.cell(row=4, column=3, value=4.2)
    ws4.cell(row=4, column=3).fill = LIGHT_YELLOW
    ws4.cell(row=4, column=3).border = THIN_BORDER
    ws4.cell(row=5, column=3, value=72)
    ws4.cell(row=5, column=3).fill = LIGHT_YELLOW
    ws4.cell(row=5, column=3).border = THIN_BORDER

    ws4.cell(row=6, column=3).value = "=ROUND(C4/5*100*0.7,1)"
    ws4.cell(row=7, column=3).value = "=ROUND(C5*0.3,1)"
    ws4.cell(row=8, column=3).value = "=C6+C7"
    ws4.cell(row=8, column=3).font = Font(bold=True, size=14, color="1F4E79")
    ws4.cell(row=10, column=3).value = '=IF(C8>=85,"🏆 MASTER",IF(C8>=70,"⭐ EXEMPLARY",IF(C8>=55,"✅ RECOGNIZED","Not Yet Designated")))'
    ws4.cell(row=10, column=3).font = Font(bold=True, size=14)

    # Reference table
    ws4.cell(row=13, column=1, value="DESIGNATION REFERENCE TABLE").font = INST_BOLD
    ref_headers = ["Designation", "Composite Score Range", "Annual Stipend", "Example: Obs Avg", "Example: Growth Score"]
    for c, h in enumerate(ref_headers, 1):
        ws4.cell(row=14, column=c, value=h)
    style_header(ws4, 14, len(ref_headers))

    ref_data = [
        ["Recognized", "55 - 69", "$3,000 - $6,000", "3.5", "55"],
        ["Exemplary", "70 - 84", "$6,000 - $12,000", "4.0", "70"],
        ["Master", "85 - 100", "$12,000 - $32,000", "4.5", "85"],
    ]
    for i, row in enumerate(ref_data):
        r = 15 + i
        for c, val in enumerate(row, 1):
            ws4.cell(row=r, column=c, value=val)
        fills = [LIGHT_GREEN, LIGHT_YELLOW, PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")]
        for c in range(1, 6):
            ws4.cell(row=r, column=c).fill = fills[i]
            ws4.cell(row=r, column=c).border = THIN_BORDER

    auto_width(ws4, 5, min_w=18)
    ws4.column_dimensions["A"].width = 38

    # ── Instructions ──
    ws5 = wb.create_sheet("Instructions")
    write_instructions(ws5, [
        ("📋 TIA Portfolio Builder — Instructions", True),
        ("", False),
        ("WHAT IS TIA?", True),
        ("The Teacher Incentive Allotment (TIA) is a Texas program that rewards effective teachers", False),
        ("with annual stipends of $3,000 to $32,000+ based on their designation level.", False),
        ("Designations: Recognized → Exemplary → Master", False),
        ("TIA scores are based on Teacher Observation (70%) + Student Growth (30%)", False),
        ("", False),
        ("HOW THIS TRACKER HELPS", True),
        ("• Observation Tracker: Log every formal/informal observation with scores and evidence", False),
        ("• Student Growth: Track assessment data that feeds into your TIA growth score", False),
        ("• Professional Development: Document PD hours aligned to T-TESS domains", False),
        ("• TIA Calculator: Input your scores to see your projected designation", False),
        ("", False),
        ("OBSERVATION TRACKER TAB", True),
        ("• Log every observation immediately — don't wait until end of year", False),
        ("• Include specific evidence (student work samples, lesson artifacts)", False),
        ("• Note which T-TESS dimension each observation addresses", False),
        ("• Scores: 1=Improvement Needed, 2=Developing, 3=Proficient, 4=Accomplished, 5=Distinguished", False),
        ("", False),
        ("STUDENT GROWTH TAB", True),
        ("• Enter baseline (BOY) and target scores for each assessment", False),
        ("• Update actual scores as results come in", False),
        ("• Growth % and target status auto-calculate", False),
        ("• Top Quartile growth = strongest TIA evidence", False),
        ("", False),
        ("TIA CALCULATOR TAB", True),
        ("• Yellow cells = your inputs (observation avg and growth score)", False),
        ("• Everything else calculates automatically", False),
        ("• Reference table shows what each designation level looks like", False),
        ("• Use this to set goals: 'I need a 4.2 observation avg and 70 growth to hit Exemplary'", False),
        ("", False),
        ("TIPS FOR BUILDING YOUR PORTFOLIO", True),
        ("• Start documenting NOW — don't wait until TIA application opens", False),
        ("• Take photos of anchor charts, student work, and classroom setup", False),
        ("• Save emails from parents praising your teaching", False),
        ("• Keep PD certificates in a Google Drive folder", False),
        ("• Ask your appraiser for specific feedback on each dimension", False),
        ("", False),
        ("© teach4texas | For personal classroom use only", False),
    ])

    wb.save(os.path.join(OUT, "TIA_Portfolio_Builder.xlsx"))
    print("✅ TIA_Portfolio_Builder.xlsx")


# ════════════════════════════════════════════════════════════════════════
# 3. Teacher Budget Planner
# ════════════════════════════════════════════════════════════════════════

def build_budget():
    wb = openpyxl.Workbook()

    # ── Monthly Budget ──
    ws = wb.active
    ws.title = "Monthly Budget"
    ws.sheet_properties.tabColor = "548235"

    ws.cell(row=1, column=1, value="MONTHLY INCOME & EXPENSES").font = TITLE_FONT
    ws.merge_cells("A1:D1")

    # Income section
    ws.cell(row=3, column=1, value="INCOME").font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=3, column=1).fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    for c in range(1, 5):
        ws.cell(row=3, column=c).fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        ws.cell(row=3, column=c).font = Font(bold=True, color="FFFFFF")
    headers_inc = ["Source", "Monthly Amount", "Annual Amount", "Notes"]
    for c, h in enumerate(headers_inc, 1):
        ws.cell(row=3, column=c, value=h)

    income_rows = [
        ["Base Salary (after taxes)", 3800, "=B4*12", "Gross ~$55K, net ~$45.6K"],
        ["Tutoring", 400, "=B5*12", "2 students × $50/hr × 4 sessions"],
        ["TPT Sales", 150, "=B6*12", "Average monthly — varies seasonally"],
        ["Side Hustle (other)", 0, "=B7*12", "Curriculum writing, workshops, etc."],
        ["TIA Stipend (monthly)", 500, "=B8*12", "Exemplary = ~$6K-$12K/yr"],
    ]
    for i, row in enumerate(income_rows):
        r = 4 + i
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
            ws.cell(row=r, column=c).border = THIN_BORDER
        ws.cell(row=r, column=2).number_format = '$#,##0'
        ws.cell(row=r, column=3).number_format = '$#,##0'
        if i % 2 == 1:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = ALT_ROW

    r_inc_total = 9
    ws.cell(row=r_inc_total, column=1, value="TOTAL INCOME").font = Font(bold=True)
    ws.cell(row=r_inc_total, column=2).value = "=SUM(B4:B8)"
    ws.cell(row=r_inc_total, column=2).font = Font(bold=True, color="548235")
    ws.cell(row=r_inc_total, column=2).number_format = '$#,##0'
    for c in range(1, 5):
        ws.cell(row=r_inc_total, column=c).border = THIN_BORDER

    # Expenses section
    ws.cell(row=11, column=1, value="EXPENSES").font = Font(bold=True, size=12, color="FFFFFF")
    for c in range(1, 5):
        ws.cell(row=11, column=c).fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        ws.cell(row=11, column=c).font = Font(bold=True, color="FFFFFF")
    headers_exp = ["Category", "Monthly Amount", "Annual Amount", "Notes"]
    for c, h in enumerate(headers_exp, 1):
        ws.cell(row=11, column=c, value=h)

    expense_rows = [
        ["Rent/Mortgage", 1200, "=B12*12", ""],
        ["Car Payment", 350, "=B13*12", ""],
        ["Car Insurance", 150, "=B14*12", ""],
        ["Gas/Transportation", 120, "=B15*12", ""],
        ["Student Loans", 300, "=B16*12", "PSLF eligible — 10 yr forgiveness"],
        ["Groceries", 400, "=B17*12", ""],
        ["Utilities (electric, water, internet)", 200, "=B18*12", ""],
        ["Phone", 80, "=B19*12", ""],
        ["Health Insurance", 150, "=B20*12", "District plan"],
        ["Subscriptions (Netflix, Spotify, etc.)", 50, "=B21*12", ""],
        ["Dining Out / Coffee", 150, "=B22*12", ""],
        ["Classroom Spending", 75, "=B23*12", "See Classroom Spending tab"],
        ["Savings", 200, "=B24*12", "Emergency fund goal"],
        ["Other / Misc", 100, "=B25*12", ""],
    ]
    for i, row in enumerate(expense_rows):
        r = 12 + i
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
            ws.cell(row=r, column=c).border = THIN_BORDER
        ws.cell(row=r, column=2).number_format = '$#,##0'
        ws.cell(row=r, column=3).number_format = '$#,##0'
        if i % 2 == 1:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = ALT_ROW

    r_exp_total = 26
    ws.cell(row=r_exp_total, column=1, value="TOTAL EXPENSES").font = Font(bold=True)
    ws.cell(row=r_exp_total, column=2).value = "=SUM(B12:B25)"
    ws.cell(row=r_exp_total, column=2).font = Font(bold=True, color="C00000")
    ws.cell(row=r_exp_total, column=2).number_format = '$#,##0'
    for c in range(1, 5):
        ws.cell(row=r_exp_total, column=c).border = THIN_BORDER

    # Net
    ws.cell(row=28, column=1, value="NET MONTHLY (Income - Expenses)").font = Font(bold=True, size=13)
    ws.cell(row=28, column=2).value = "=B9-B26"
    ws.cell(row=28, column=2).font = Font(bold=True, size=13, color="1F4E79")
    ws.cell(row=28, column=2).number_format = '$#,##0'

    auto_width(ws, 4, min_w=18)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["D"].width = 35

    # ── Classroom Spending ──
    ws2 = wb.create_sheet("Classroom Spending")
    ws2.sheet_properties.tabColor = "BF8F00"
    headers2 = ["Date", "Item", "Category", "Cost", "Reimbursed?", "Out of Pocket", "Notes"]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, len(headers2))

    spend_data = [
        ["2025-08-10", "Expo markers (bulk)", "Supplies", 24.99, "No", None, "Back to school"],
        ["2025-08-10", "Bulletin board borders", "Decor", 18.50, "No", None, "Classroom setup"],
        ["2025-08-15", "Composition notebooks (30-pk)", "Supplies", 32.00, "Yes", None, "Reimbursed by campus"],
        ["2025-08-20", "Flexible seating cushions", "Decor", 45.00, "No", None, "Amazon"],
        ["2025-09-05", "Laminating pouches", "Supplies", 12.99, "No", None, ""],
        ["2025-09-15", "Kahoot premium subscription", "Tech", 48.00, "Yes", None, "District paid"],
        ["2025-10-01", "Halloween incentive prizes", "Supplies", 15.00, "No", None, "Dollar Tree"],
        ["2025-10-20", "Scholastic book order (class set)", "Curriculum", 35.00, "No", None, "Used bonus points too"],
        ["2025-11-15", "Chart paper (10-pk)", "Supplies", 22.00, "No", None, ""],
        ["2025-12-10", "Holiday party supplies", "Supplies", 20.00, "No", None, "Shared cost with team"],
        ["2026-01-05", "New pencil sharpener", "Supplies", 28.99, "No", None, "Electric — worth it"],
        ["2026-01-20", "STAAR prep workbooks", "Curriculum", 55.00, "Yes", None, "Title I funds"],
    ]
    for i, row in enumerate(spend_data):
        r = i + 2
        for c, val in enumerate(row, 1):
            if val is not None:
                ws2.cell(row=r, column=c, value=val)
        ws2.cell(row=r, column=4).number_format = '$#,##0.00'
        ws2.cell(row=r, column=6).value = f'=IF(E{r}="Yes",0,D{r})'
        ws2.cell(row=r, column=6).number_format = '$#,##0.00'
    style_data_rows(ws2, 2, len(spend_data)+1, len(headers2))

    # Totals
    tr = len(spend_data) + 3
    ws2.cell(row=tr, column=3, value="TOTALS:").font = Font(bold=True)
    ws2.cell(row=tr, column=4).value = f"=SUM(D2:D{len(spend_data)+1})"
    ws2.cell(row=tr, column=4).font = Font(bold=True)
    ws2.cell(row=tr, column=4).number_format = '$#,##0.00'
    ws2.cell(row=tr, column=6).value = f"=SUM(F2:F{len(spend_data)+1})"
    ws2.cell(row=tr, column=6).font = Font(bold=True, color="C00000")
    ws2.cell(row=tr, column=6).number_format = '$#,##0.00'

    # Category breakdown
    tr2 = tr + 2
    ws2.cell(row=tr2, column=1, value="SPENDING BY CATEGORY").font = INST_BOLD
    cats = ["Supplies", "Decor", "Tech", "Curriculum"]
    for i, cat in enumerate(cats):
        r = tr2 + 1 + i
        ws2.cell(row=r, column=1, value=cat)
        ws2.cell(row=r, column=2).value = f'=SUMIF(C2:C{len(spend_data)+1},A{r},D2:D{len(spend_data)+1})'
        ws2.cell(row=r, column=2).number_format = '$#,##0.00'

    auto_width(ws2, len(headers2))
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["G"].width = 25

    # ── Summer Income Planner ──
    ws3 = wb.create_sheet("Summer Planner")
    ws3.sheet_properties.tabColor = "ED7D31"
    ws3.cell(row=1, column=1, value="SUMMER INCOME PLANNER").font = TITLE_FONT
    ws3.merge_cells("A1:D1")

    headers3a = ["Income Source", "June", "July", "Total"]
    for c, h in enumerate(headers3a, 1):
        ws3.cell(row=3, column=c, value=h)
    style_header(ws3, 3, len(headers3a))

    summer_inc = [
        ["Summer School Teaching", 2500, 2500, "=B4+C4"],
        ["Private Tutoring", 800, 600, "=B5+C5"],
        ["TPT Sales", 200, 300, "=B6+C6"],
        ["Camp Counselor / Other", 0, 0, "=B7+C7"],
        ["Curriculum Writing", 500, 500, "=B8+C8"],
    ]
    for i, row in enumerate(summer_inc):
        r = 4 + i
        for c, val in enumerate(row, 1):
            ws3.cell(row=r, column=c, value=val)
            ws3.cell(row=r, column=c).border = THIN_BORDER
        for c in [2,3,4]:
            ws3.cell(row=r, column=c).number_format = '$#,##0'
        if i % 2 == 1:
            for c in range(1, 5):
                ws3.cell(row=r, column=c).fill = ALT_ROW

    ws3.cell(row=9, column=1, value="TOTAL SUMMER INCOME").font = Font(bold=True)
    ws3.cell(row=9, column=4).value = "=SUM(D4:D8)"
    ws3.cell(row=9, column=4).font = Font(bold=True, color="548235")
    ws3.cell(row=9, column=4).number_format = '$#,##0'

    # Savings goals
    ws3.cell(row=11, column=1, value="SUMMER SAVINGS GOALS").font = INST_BOLD
    goals = [
        ["Emergency Fund Contribution", 1000],
        ["Back-to-School Classroom Setup", 300],
        ["Vacation / Self-Care", 500],
        ["Professional Development", 200],
    ]
    goal_headers = ["Goal", "Target Amount", "Saved So Far", "Remaining"]
    for c, h in enumerate(goal_headers, 1):
        ws3.cell(row=12, column=c, value=h)
    style_header(ws3, 12, len(goal_headers))
    for i, (goal, amt) in enumerate(goals):
        r = 13 + i
        ws3.cell(row=r, column=1, value=goal)
        ws3.cell(row=r, column=2, value=amt)
        ws3.cell(row=r, column=2).number_format = '$#,##0'
        ws3.cell(row=r, column=3, value=0)
        ws3.cell(row=r, column=3).number_format = '$#,##0'
        ws3.cell(row=r, column=3).fill = LIGHT_YELLOW
        ws3.cell(row=r, column=4).value = f"=B{r}-C{r}"
        ws3.cell(row=r, column=4).number_format = '$#,##0'
        for c in range(1, 5):
            ws3.cell(row=r, column=c).border = THIN_BORDER

    auto_width(ws3, 4, min_w=18)
    ws3.column_dimensions["A"].width = 30

    # ── Salary Comparison ──
    ws4 = wb.create_sheet("Salary Comparison")
    ws4.sheet_properties.tabColor = "1F4E79"
    ws4.cell(row=1, column=1, value="SALARY & COMPENSATION COMPARISON").font = TITLE_FONT
    ws4.merge_cells("A1:F1")

    headers4 = ["Component", "Current District", "District B", "District C", "Notes"]
    for c, h in enumerate(headers4, 1):
        ws4.cell(row=3, column=c, value=h)
    style_header(ws4, 3, len(headers4))

    comp_data = [
        ["Base Salary", 55000, 58000, 52000, "Check salary schedule for your step"],
        ["Bilingual Stipend", 3000, 2500, 4000, "If applicable"],
        ["GT Stipend", 1500, 1000, 2000, "If applicable"],
        ["Master's Degree Stipend", 1500, 2000, 1000, ""],
        ["TIA - Recognized", 5000, 3000, 6000, "Varies by district allotment"],
        ["TIA - Exemplary", 9000, 7000, 12000, ""],
        ["TIA - Master", 20000, 15000, 32000, "Rural districts often pay more"],
        ["Health Insurance (district share)", 0, 0, 0, "Check out-of-pocket difference"],
        ["TOTAL (with Exemplary TIA)", None, None, None, ""],
    ]
    for i, row in enumerate(comp_data):
        r = 4 + i
        for c, val in enumerate(row, 1):
            if val is not None:
                ws4.cell(row=r, column=c, value=val)
            ws4.cell(row=r, column=c).border = THIN_BORDER
        for c in [2,3,4]:
            if row[c-1] is not None and isinstance(row[c-1], (int, float)):
                ws4.cell(row=r, column=c).number_format = '$#,##0'
        if i % 2 == 1:
            for c in range(1, 6):
                ws4.cell(row=r, column=c).fill = ALT_ROW

    # Total row (sum base + exemplary TIA + applicable stipends)
    total_r = 12  # row for TOTAL
    for c in [2, 3, 4]:
        col_l = get_column_letter(c)
        ws4.cell(row=total_r, column=c).value = f"={col_l}4+{col_l}9"  # base + exemplary
        ws4.cell(row=total_r, column=c).number_format = '$#,##0'
        ws4.cell(row=total_r, column=c).font = Font(bold=True, size=12)

    auto_width(ws4, 5, min_w=18)
    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["E"].width = 35

    # ── Instructions ──
    ws5 = wb.create_sheet("Instructions")
    write_instructions(ws5, [
        ("📋 Teacher Budget Planner — Instructions", True),
        ("", False),
        ("MONTHLY BUDGET TAB", True),
        ("• Update the yellow/editable amounts to match YOUR actual income and expenses", False),
        ("• Annual amounts auto-calculate (monthly × 12)", False),
        ("• Net Monthly shows what's left after all expenses", False),
        ("• Include ALL income sources: salary, tutoring, TPT, TIA stipend", False),
        ("• Remember: Texas has no state income tax! 🤠", False),
        ("", False),
        ("CLASSROOM SPENDING TAB", True),
        ("• Log every classroom purchase — you'll be shocked how much it adds up", False),
        ("• Mark whether it was reimbursed (campus funds, Title I, PTA, etc.)", False),
        ("• Out of Pocket auto-calculates for non-reimbursed items", False),
        ("• Category breakdown shows where your money goes", False),
        ("• Tax tip: Teachers can deduct up to $300 of unreimbursed classroom expenses", False),
        ("", False),
        ("SUMMER PLANNER TAB", True),
        ("• Plan your summer income BEFORE school ends", False),
        ("• Set specific savings goals and track progress", False),
        ("• Common summer income: summer school, tutoring, curriculum writing, camps", False),
        ("", False),
        ("SALARY COMPARISON TAB", True),
        ("• Compare total compensation across districts — not just base salary", False),
        ("• Don't forget TIA stipends — they can be $3K-$32K+ difference!", False),
        ("• Check if districts pay TIA differently (some spread monthly, some lump sum)", False),
        ("• Factor in commute costs, health insurance differences, and stipends you qualify for", False),
        ("", False),
        ("TEACHER MONEY TIPS", True),
        ("• Explore PSLF (Public Service Loan Forgiveness) for student loans — you qualify!", False),
        ("• Max out your 403(b) match if your district offers one", False),
        ("• Track classroom spending for tax deductions ($300 educator expense deduction)", False),
        ("• TIA designation = the biggest raise you can give yourself", False),
        ("", False),
        ("© teach4texas | For personal classroom use only", False),
    ])

    wb.save(os.path.join(OUT, "Teacher_Budget_Planner.xlsx"))
    print("✅ Teacher_Budget_Planner.xlsx")


if __name__ == "__main__":
    build_staar()
    build_tia()
    build_budget()
    print("\n🎉 All 3 templates built!")
