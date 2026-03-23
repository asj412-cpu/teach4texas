#!/usr/bin/env python3
"""Build Student Growth Tracking Template as .xlsx"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, LineChart, BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

# Colors
BLUE = "4472C4"
WHITE = "FFFFFF"
LIGHT_GREEN = "C6EFCE"
LIGHT_YELLOW = "FFEB9C"
LIGHT_RED = "FFC7CE"
LIGHT_BLUE = "BDD7EE"
DARK_GREEN = "548235"
MED_GREEN = "A9D18E"
YELLOW_APPROACH = "FFD966"
RED_DNM = "FF6B6B"
LIGHT_GRAY = "F2F2F2"

header_font = Font(bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
header_align = Alignment(horizontal="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

def style_range(ws, min_row, max_row, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if c > 3 else "left")

MAX_STUDENTS = 35
DATA_START = 3  # row 3 (row 1=header, row 2=descriptions)
DATA_END = DATA_START + MAX_STUDENTS - 1  # row 37

wb = openpyxl.Workbook()

# ============================================================
# TAB 1: Class Roster
# ============================================================
ws1 = wb.active
ws1.title = "Class Roster"
ws1.sheet_properties.tabColor = BLUE

headers1 = ["Last Name", "First Name", "Student ID", "Subgroup Tags", "Grade Level", "Notes"]
widths1 = [20, 20, 15, 30, 12, 35]

for i, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=i, value=h)
    ws1.column_dimensions[get_column_letter(i)].width = widths1[i-1]

# Row 2: descriptions
descs1 = ["Required", "Required", "Required", "ELL, SpEd, GT, 504, Dyslexia, At-Risk, Bilingual", "Optional", "Optional"]
desc_font = Font(italic=True, color="808080", size=9)
for i, d in enumerate(descs1, 1):
    cell = ws1.cell(row=2, column=i, value=d)
    cell.font = desc_font

style_header(ws1, 1, 6)
ws1.freeze_panes = "A3"

# Data validation for subgroups
dv = DataValidation(type="list", formula1='"ELL,SpEd,GT,504,Dyslexia,At-Risk,Bilingual"', allow_blank=True)
dv.prompt = "Select subgroup(s)"
ws1.add_data_validation(dv)
dv.add(f"D{DATA_START}:D{DATA_END}")

# Add sample data (3 students)
samples = [
    ["Smith", "John", "12345", "ELL", 4, ""],
    ["Johnson", "Maria", "12346", "GT", 4, ""],
    ["Williams", "Carlos", "12347", "ELL,At-Risk", 4, "Transfer student"],
]
for r, row_data in enumerate(samples, DATA_START):
    for c, val in enumerate(row_data, 1):
        ws1.cell(row=r, column=c, value=val)

style_range(ws1, DATA_START, DATA_START + 2, 6)

# ============================================================
# TAB 2: Assessment Scores
# ============================================================
ws2 = wb.create_sheet("Assessment Scores")
ws2.sheet_properties.tabColor = "2E75B6"

headers2 = ["Last Name", "First Name", "Student ID",
            "Pre-Test", "Unit 1", "Unit 2", "Unit 3", "Unit 4",
            "Unit 5", "Unit 6", "Unit 7", "Unit 8",
            "Mid-Year", "Post-Test",
            "STAAR Mock 1", "STAAR Mock 2", "STAAR Mock 3",
            "Current Average", "Trend"]
widths2 = [20, 20, 15] + [12]*14 + [15, 12]

for i, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=i, value=h)
    ws2.column_dimensions[get_column_letter(i)].width = widths2[i-1]

style_header(ws2, 1, len(headers2))
ws2.freeze_panes = "D3"

# Row 2: descriptions
descs2 = ["(auto)", "(auto)", "(auto)"] + ["0-100"]*14 + ["(auto)", "(auto)"]
for i, d in enumerate(descs2, 1):
    cell = ws2.cell(row=2, column=i, value=d)
    cell.font = desc_font

# Auto-populate and formulas for each student row
for r in range(DATA_START, DATA_END + 1):
    # Auto-populate names from roster
    ws2.cell(row=r, column=1, value=f"='Class Roster'!A{r}")
    ws2.cell(row=r, column=2, value=f"='Class Roster'!B{r}")
    ws2.cell(row=r, column=3, value=f"='Class Roster'!C{r}")
    # Current Average (col R = 18)
    ws2.cell(row=r, column=18, value=f'=IFERROR(AVERAGE(D{r}:Q{r}),"")')
    # Trend (col S = 19)
    ws2.cell(row=r, column=19, value=f'=IF(AND(D{r}<>"",N{r}<>""),IF(N{r}>D{r},"↑ Rising",IF(N{r}<D{r},"↓ Falling","→ Flat")),"")')

# Sample scores
sample_scores = [
    [45, 52, 58, 60, 63, 55, 62, 68, 70, 65, 72, 55, 60, 65],
    [82, 85, 88, 90, 87, 92, 89, 91, 93, 88, 95, 87, 90, 92],
    [38, 42, 45, 48, 44, 40, 47, 50, 52, 45, 48, 43, 46, 50],
]
for r, scores in enumerate(sample_scores, DATA_START):
    for c, score in enumerate(scores, 4):
        ws2.cell(row=r, column=c, value=score)

# Conditional formatting for scores
from openpyxl.formatting.rule import CellIsRule
red_fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
yellow_fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
blue_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
green_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")

score_range = f"D{DATA_START}:Q{DATA_END}"
ws2.conditional_formatting.add(score_range, CellIsRule(operator='lessThan', formula=['50'], fill=red_fill))
ws2.conditional_formatting.add(score_range, CellIsRule(operator='between', formula=['50', '69'], fill=yellow_fill))
ws2.conditional_formatting.add(score_range, CellIsRule(operator='between', formula=['70', '84'], fill=blue_fill))
ws2.conditional_formatting.add(score_range, CellIsRule(operator='between', formula=['85', '100'], fill=green_fill))

style_range(ws2, DATA_START, DATA_START + 2, len(headers2))

# ============================================================
# TAB 3: Growth Analysis
# ============================================================
ws3 = wb.create_sheet("Growth Analysis")
ws3.sheet_properties.tabColor = "2E8B57"

headers3 = ["Last Name", "First Name", "Subgroup", "Pre-Test Score",
            "Most Recent Score", "Point Growth", "% Growth",
            "Growth Rating", "Pre→Mid Growth", "Mid→Post Growth",
            "STAAR Mock Trend", "Intervention Flag"]
widths3 = [20, 20, 18, 14, 14, 14, 12, 20, 14, 14, 16, 22]

for i, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=i, value=h)
    ws3.column_dimensions[get_column_letter(i)].width = widths3[i-1]

style_header(ws3, 1, len(headers3))
ws3.freeze_panes = "D2"

for r in range(DATA_START, DATA_END + 1):
    ws3.cell(row=r, column=1, value=f"='Class Roster'!A{r}")
    ws3.cell(row=r, column=2, value=f"='Class Roster'!B{r}")
    ws3.cell(row=r, column=3, value=f"='Class Roster'!D{r}")
    ws3.cell(row=r, column=4, value=f"='Assessment Scores'!D{r}")
    # Most Recent Score - simplified: use column R (average) as proxy
    ws3.cell(row=r, column=5, value=f"='Assessment Scores'!R{r}")
    # Point Growth
    ws3.cell(row=r, column=6, value=f'=IFERROR(E{r}-D{r},"")')
    # % Growth
    ws3.cell(row=r, column=7, value=f'=IFERROR(IF(D{r}=0,"",(E{r}-D{r})/D{r}*100),"")')
    # Growth Rating
    ws3.cell(row=r, column=8, value=f'=IF(G{r}="","",IF(G{r}<10,"🔴 Low Growth",IF(G{r}<20,"🟡 Moderate","🟢 Strong Growth")))')
    # Pre→Mid Growth
    ws3.cell(row=r, column=9, value=f"=IFERROR('Assessment Scores'!M{r}-'Assessment Scores'!D{r},\"\")")
    # Mid→Post Growth
    ws3.cell(row=r, column=10, value=f"=IFERROR('Assessment Scores'!N{r}-'Assessment Scores'!M{r},\"\")")
    # STAAR Mock Trend
    ws3.cell(row=r, column=11, value=f"=IF(AND('Assessment Scores'!O{r}<>\"\",\"\"<>'Assessment Scores'!P{r}),IF('Assessment Scores'!P{r}>'Assessment Scores'!O{r},\"↑\",\"↓\"),\"\")")
    # Intervention Flag
    ws3.cell(row=r, column=12, value=f'=IF(OR(G{r}<10,E{r}<50),"⚠️ NEEDS INTERVENTION","")')

# Conditional formatting for % Growth column (G)
growth_range = f"G{DATA_START}:G{DATA_END}"
ws3.conditional_formatting.add(growth_range, CellIsRule(operator='lessThan', formula=['10'], fill=red_fill))
ws3.conditional_formatting.add(growth_range, CellIsRule(operator='between', formula=['10', '20'], fill=yellow_fill))
ws3.conditional_formatting.add(growth_range, CellIsRule(operator='greaterThan', formula=['20'], fill=green_fill))

# ============================================================
# TAB 4: STAAR Readiness
# ============================================================
ws4 = wb.create_sheet("STAAR Readiness")
ws4.sheet_properties.tabColor = "FFC000"

headers4 = ["Last Name", "First Name", "Subgroup",
            "BOY Level", "MOY Level", "Mock 1 Level", "Mock 2 Level",
            "Current Projection", "Movement", "Action Needed"]
widths4 = [20, 20, 18, 12, 12, 14, 14, 16, 18, 45]

for i, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=i, value=h)
    ws4.column_dimensions[get_column_letter(i)].width = widths4[i-1]

style_header(ws4, 1, len(headers4))
ws4.freeze_panes = "D2"

# Data validation for D/A/M/MA
dv_staar = DataValidation(type="list", formula1='"D,A,M,MA"', allow_blank=True)
dv_staar.prompt = "D=Did Not Meet, A=Approaches, M=Meets, MA=Masters"
ws4.add_data_validation(dv_staar)
dv_staar.add(f"D{DATA_START}:G{DATA_END}")

for r in range(DATA_START, DATA_END + 1):
    ws4.cell(row=r, column=1, value=f"='Class Roster'!A{r}")
    ws4.cell(row=r, column=2, value=f"='Class Roster'!B{r}")
    ws4.cell(row=r, column=3, value=f"='Class Roster'!D{r}")
    # Current Projection (most recent non-blank in D-G)
    ws4.cell(row=r, column=8, value=f'=IFERROR(LOOKUP(2,1/(D{r}:G{r}<>""),D{r}:G{r}),"")')
    # Movement
    ws4.cell(row=r, column=9, value=(
        f'=IF(AND(D{r}<>"",H{r}<>""),'
        f'IF(MATCH(H{r},{{"D","A","M","MA"}},0)>MATCH(D{r},{{"D","A","M","MA"}},0),"⬆️ Moving Up",'
        f'IF(MATCH(H{r},{{"D","A","M","MA"}},0)<MATCH(D{r},{{"D","A","M","MA"}},0),"⬇️ Moving Down","➡️ No Change")),"")'
    ))
    # Action Needed
    ws4.cell(row=r, column=10, value=(
        f'=IF(H{r}="D","🚨 Intensive Intervention",'
        f'IF(H{r}="A","📋 Targeted Practice",'
        f'IF(H{r}="M","💪 Push toward Masters",'
        f'IF(H{r}="MA","⭐ Extension & peer tutoring",""))))'
    ))

# Sample data
staar_samples = [
    ["D", "A", "A", "A"],
    ["M", "M", "MA", "MA"],
    ["D", "D", "D", "A"],
]
for r, levels in enumerate(staar_samples, DATA_START):
    for c, level in enumerate(levels, 4):
        ws4.cell(row=r, column=c, value=level)

# Conditional formatting for STAAR levels
from openpyxl.formatting.rule import FormulaRule
dnm_fill = PatternFill(start_color=RED_DNM, end_color=RED_DNM, fill_type="solid")
app_fill = PatternFill(start_color=YELLOW_APPROACH, end_color=YELLOW_APPROACH, fill_type="solid")
meets_fill = PatternFill(start_color=MED_GREEN, end_color=MED_GREEN, fill_type="solid")
masters_fill = PatternFill(start_color=DARK_GREEN, end_color=DARK_GREEN, fill_type="solid")

staar_range = f"D{DATA_START}:G{DATA_END}"
ws4.conditional_formatting.add(staar_range, FormulaRule(formula=[f'D{DATA_START}="D"'], fill=dnm_fill))
ws4.conditional_formatting.add(staar_range, FormulaRule(formula=[f'D{DATA_START}="A"'], fill=app_fill))
ws4.conditional_formatting.add(staar_range, FormulaRule(formula=[f'D{DATA_START}="M"'], fill=meets_fill))
ws4.conditional_formatting.add(staar_range, FormulaRule(formula=[f'D{DATA_START}="MA"'], fill=masters_fill))

# ============================================================
# TAB 5: Dashboard
# ============================================================
ws5 = wb.create_sheet("Dashboard")
ws5.sheet_properties.tabColor = "FF4500"

title_font = Font(bold=True, size=18, color=BLUE)
section_font = Font(bold=True, size=14, color=BLUE)
label_font = Font(bold=True, size=11)
value_font = Font(size=11)

# Title
ws5.merge_cells("A1:F1")
ws5.cell(row=1, column=1, value="📊 CLASS DASHBOARD").font = title_font

# Class Overview
ws5.cell(row=3, column=1, value="CLASS OVERVIEW").font = section_font
labels = [
    ("Total Students:", f"=COUNTA('Class Roster'!A{DATA_START}:A{DATA_END})"),
    ("Class Pre-Test Average:", f"=IFERROR(AVERAGE('Assessment Scores'!D{DATA_START}:D{DATA_END}),\"\")"),
    ("Class Current Average:", f"=IFERROR(AVERAGE('Growth Analysis'!E{DATA_START}:E{DATA_END}),\"\")"),
    ("Class Average Growth (%):", f"=IFERROR(AVERAGE('Growth Analysis'!G{DATA_START}:G{DATA_END}),\"\")"),
    ("Students Needing Intervention:", f'=COUNTIF(\'Growth Analysis\'!L{DATA_START}:L{DATA_END},"⚠️*")'),
]
for i, (label, formula) in enumerate(labels, 4):
    ws5.cell(row=i, column=1, value=label).font = label_font
    ws5.cell(row=i, column=2, value=formula).font = value_font

# STAAR Readiness Breakdown
ws5.cell(row=10, column=1, value="STAAR READINESS BREAKDOWN").font = section_font
staar_labels = [
    ("Did Not Meet (D):", f'=COUNTIF(\'STAAR Readiness\'!H{DATA_START}:H{DATA_END},"D")'),
    ("Approaches (A):", f'=COUNTIF(\'STAAR Readiness\'!H{DATA_START}:H{DATA_END},"A")'),
    ("Meets (M):", f'=COUNTIF(\'STAAR Readiness\'!H{DATA_START}:H{DATA_END},"M")'),
    ("Masters (MA):", f'=COUNTIF(\'STAAR Readiness\'!H{DATA_START}:H{DATA_END},"MA")'),
]
for i, (label, formula) in enumerate(staar_labels, 11):
    ws5.cell(row=i, column=1, value=label).font = label_font
    ws5.cell(row=i, column=2, value=formula).font = value_font

# Subgroup Breakdown
ws5.cell(row=17, column=1, value="SUBGROUP GROWTH COMPARISON").font = section_font
subgroups = ["ELL", "SpEd", "GT", "504", "At-Risk"]
for i, sg in enumerate(subgroups, 18):
    ws5.cell(row=i, column=1, value=f"{sg} Avg Growth:").font = label_font
    ws5.cell(row=i, column=2, value=(
        f'=IFERROR(AVERAGEIF(\'Growth Analysis\'!C{DATA_START}:C{DATA_END},"*{sg}*",\'Growth Analysis\'!G{DATA_START}:G{DATA_END}),"")'
    )).font = value_font

# Column widths
ws5.column_dimensions["A"].width = 35
ws5.column_dimensions["B"].width = 15
ws5.column_dimensions["C"].width = 12

# Pie chart for STAAR readiness
pie = PieChart()
pie.title = "STAAR Readiness Distribution"
pie.style = 10
labels_ref = Reference(ws5, min_col=1, min_row=11, max_row=14)
data_ref = Reference(ws5, min_col=2, min_row=11, max_row=14)
pie.add_data(data_ref)
pie.set_categories(labels_ref)
pie.width = 18
pie.height = 12
ws5.add_chart(pie, "D3")

# ============================================================
# README tab
# ============================================================
ws_readme = wb.create_sheet("README", 0)  # Insert at position 0
ws_readme.sheet_properties.tabColor = "00B050"

readme_font = Font(size=11)
readme_title = Font(bold=True, size=16, color=BLUE)
readme_h2 = Font(bold=True, size=13, color=BLUE)

ws_readme.column_dimensions["A"].width = 80
ws_readme.cell(row=1, column=1, value="📊 Student Growth Tracking Template — Setup Guide").font = readme_title
ws_readme.cell(row=3, column=1, value="⚡ QUICK START (5 minutes!)").font = readme_h2

instructions = [
    "1. Go to the 'Class Roster' tab and enter your students' names, IDs, and subgroup tags.",
    "2. Go to 'Assessment Scores' — names auto-populate! Just enter scores as you get them.",
    "3. The 'Growth Analysis' tab auto-calculates everything — growth, ratings, intervention flags.",
    "4. Use the 'STAAR Readiness' tab to track D/A/M/MA levels across assessment windows.",
    "5. The 'Dashboard' tab gives you instant charts and stats for PLC meetings!",
    "",
    "💡 TIPS:",
    "• You can enter up to 35 students per class.",
    "• Subgroup tags are comma-separated (e.g., 'ELL,At-Risk').",
    "• Scores should be 0-100. Leave blank if no score yet.",
    "• STAAR levels use: D (Did Not Meet), A (Approaches), M (Meets), MA (Masters).",
    "• All formulas are protected — just enter data in the white cells!",
    "",
    "📋 TAB OVERVIEW:",
    "• Class Roster — Enter student info here (everything else auto-populates from this)",
    "• Assessment Scores — Enter test scores (Pre-Test through STAAR Mocks)",
    "• Growth Analysis — AUTO: Point growth, % growth, ratings, intervention flags",
    "• STAAR Readiness — Track D/A/M/MA levels and see movement trends",
    "• Dashboard — AUTO: Charts, stats, and intervention counts",
    "",
    "❓ Questions? Message us on TPT — we respond within 24 hours!",
    "",
    "© 2025. Licensed for single-classroom use.",
]
for i, line in enumerate(instructions, 4):
    ws_readme.cell(row=i, column=1, value=line).font = readme_font

# Move Class Roster to position 1
wb.move_sheet("Class Roster", offset=0)

# Save
output = "/Users/andrewjohnson/.openclaw/workspace/projects/teacher-niche/tpt-products/student-growth-tracker/Student-Growth-Tracker.xlsx"
wb.save(output)
print(f"✅ Saved to {output}")
print(f"   Tabs: {[ws.title for ws in wb.worksheets]}")
